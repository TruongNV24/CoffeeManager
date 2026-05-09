from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelExportService:
    def ExportMonthlyReport(self, month, year, statistics, daily_rows, output_dir=None):
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        sheet_name = f"BaoCaoThang_{int(month)}"
        worksheet = workbook.create_sheet(title=sheet_name)
        self.BuildRevenueWorksheet(worksheet, month, year, statistics, daily_rows)

        target_dir = Path(output_dir) if output_dir else Path.home() / "Documents"
        target_dir.mkdir(parents=True, exist_ok=True)
        file_path = target_dir / f"BaoCao_Thang_{int(month)}_{int(year)}.xlsx"
        workbook.save(file_path)
        return str(file_path)

    def BuildRevenueWorksheet(self, ws, month, year, statistics, daily_rows):
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F3E8DC")
        title_fill = PatternFill("solid", fgColor="EED8C4")

        ws.merge_cells("A1:D1")
        ws["A1"] = f"BÁO CÁO DOANH THU THÁNG {int(month)}"
        ws["A1"].font = Font(bold=True, size=18, color="5D4037")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = title_fill

        ws["A3"] = "Tháng"
        ws["B3"] = int(month)
        ws["A4"] = "Năm"
        ws["B4"] = int(year)
        ws["A5"] = "Ngày xuất file"
        ws["B5"] = datetime.now().strftime("%d/%m/%Y %H:%M")

        ws["A7"] = "THỐNG KÊ TỔNG QUAN"
        ws["A7"].font = Font(bold=True, color="5D4037")

        overview = [
            ("Tổng doanh thu", statistics["revenue"]),
            ("Tổng hóa đơn", statistics["invoice_count"]),
            ("Tổng món bán", statistics["total_items"]),
            ("Bàn hoạt động nhiều nhất", statistics["top_table"]),
        ]
        row_idx = 8
        for label, value in overview:
            ws[f"A{row_idx}"] = label
            ws[f"B{row_idx}"] = value
            ws[f"A{row_idx}"].font = Font(bold=True)
            ws[f"A{row_idx}"].fill = header_fill
            ws[f"A{row_idx}"].border = border
            ws[f"B{row_idx}"].border = border
            row_idx += 1

        detail_start = row_idx + 2
        ws[f"A{detail_start}"] = "CHI TIẾT DOANH THU THEO NGÀY"
        ws[f"A{detail_start}"].font = Font(bold=True, color="5D4037")

        header_row = detail_start + 1
        headers = ["Ngày", "Số hóa đơn", "Số món bán", "Doanh thu"]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            cell.border = border

        current_row = header_row + 1
        for item in daily_rows:
            ws.cell(row=current_row, column=1, value=item["date"])
            ws.cell(row=current_row, column=2, value=item["invoice_count"])
            ws.cell(row=current_row, column=3, value=item["total_items"])
            ws.cell(row=current_row, column=4, value=item["revenue"])
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = border
            current_row += 1

        ws.cell(row=current_row, column=1, value="TỔNG DOANH THU THÁNG")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        total_cell = ws.cell(row=current_row, column=4, value=statistics["revenue"])
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        total_cell.font = Font(bold=True)
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).fill = header_fill
            ws.cell(row=current_row, column=col).border = border

        for row in ws.iter_rows(min_row=3, max_row=current_row, min_col=1, max_col=4):
            for cell in row:
                if cell.column == 4 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0" đ"'

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].auto_size = True
            max_len = 0
            for cell in ws[col]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col].width = max(14, min(max_len + 4, 40))
